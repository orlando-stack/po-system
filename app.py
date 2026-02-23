import streamlit as st
import pandas as pd

from utils.po_parser import parse_po_excel
from utils.normalizer import normalize_po_items

from db import ENGINE, get_db
from models import Base, POHeader, POItem

import io
from openpyxl import load_workbook


def _clean_text(x):
    if x is None:
        return ""
    return str(x).strip().replace("：", ":").replace("\u3000", " ")


def extract_po_no_from_excel(uploaded_file) -> str | None:
    """
    从 Excel 里扫出 PO No：
    找到写着“买方合同号 / PO No / PO号 ...”的格子，取右边或下边的值
    """
    try:
        data = uploaded_file.getvalue()
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        return None

    labels = {
        "买方合同号", "買方合同號",
        "po no", "po#", "po #", "pono",
        "po号", "po號",
        "po编号", "po編號",
        "purchase order", "po",
    }

    for ws in wb.worksheets:
        # 扫一个合理范围（通常 PO 表头都在前面）
        max_r = min(ws.max_row or 200, 200)
        max_c = min(ws.max_column or 40, 40)

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                text = _clean_text(v).lower()
                if not text:
                    continue

                # 允许“买方合同号:”“买方合同号	”这种
                text_norm = text.replace(":", "").replace(" ", "")
                hit = any(text_norm == lab.replace(" ", "") for lab in labels) or any(lab in text for lab in labels)

                if hit:
                    # 优先取右边
                    right = ws.cell(r, c + 1).value if c + 1 <= max_c else None
                    right2 = ws.cell(r, c + 2).value if c + 2 <= max_c else None
                    down = ws.cell(r + 1, c).value if r + 1 <= max_r else None

                    for cand in (right, right2, down):
                        cand_txt = _clean_text(cand)
                        # 过滤掉还是标签/空
                        if cand_txt and len(cand_txt) >= 4 and "合同号" not in cand_txt and "po" in cand_txt.lower():
                            return cand_txt
                        # 有些 PO 号不带 PO 前缀，也收
                        if cand_txt and len(cand_txt) >= 6 and "合同号" not in cand_txt:
                            return cand_txt

    return None


st.set_page_config(page_title="PO 系统（內部）", layout="wide")
st.title("📦 PO 系统（V1：PO 上傳 + 存庫 + 查詢）")

Base.metadata.create_all(bind=ENGINE)

tab1, tab2 = st.tabs(["1) 上傳 PO", "2) 查詢 PO"])

with tab1:
    uploaded = st.file_uploader("上傳 PO Excel（.xlsx）", type=["xlsx"])
    if uploaded:
        try:
            with st.spinner("解析中..."):
                header, items_raw = parse_po_excel(uploaded)
                items = normalize_po_items(header, items_raw)

            st.success("解析成功 ✅")
            st.subheader("PO 抬頭")
            st.json(header)

            st.subheader("明細（標準化）")
            st.dataframe(items, use_container_width=True)

            if st.button("💾 保存到資料庫", type="primary"):
                db = get_db()
                try:
                    # 先从 parser 的 header 拿
                    po_no = (
                        header.get("po_no")
                        or header.get("买方合同号")
                        or header.get("買方合同號")
                        or header.get("PO No")
                        or header.get("PO号")
                        or header.get("PO號")
                    )
                    if isinstance(po_no, str):
                        po_no = po_no.strip()

                    # 拿不到就直接扫 Excel
                    if not po_no:
                        po_no = extract_po_no_from_excel(uploaded)

                    if not po_no:
                        st.error("PO No 为空，不能保存。系统已尝试从 Excel 扫描“买方合同号/PO No”，仍未找到。")
                        st.info("下一步：我可以让页面直接显示它扫到的附近单元格，帮你确认Excel里实际位置。")
                    else:
                        h = db.get(POHeader, po_no)
                        if not h:
                            h = POHeader(
                                po_no=po_no,
                                supplier_name=header.get("supplier_name"),
                                order_date=header.get("order_date"),
                            )
                            db.add(h)
                        else:
                            h.supplier_name = header.get("supplier_name")
                            h.order_date = header.get("order_date")

                        db.query(POItem).filter(POItem.po_no == po_no).delete()

                        for _, r in items.iterrows():
                            db.add(POItem(
                                po_no=po_no,
                                seq=int(r["seq"]) if pd.notna(r["seq"]) else None,
                                product_code=str(r["product_code"]) if pd.notna(r["product_code"]) else None,
                                product_name_cn=str(r["product_name_cn"]) if pd.notna(r["product_name_cn"]) else None,
                                product_name_pt=str(r["product_name_pt"]) if pd.notna(r["product_name_pt"]) else None,
                                ncm=str(r["ncm"]) if pd.notna(r["ncm"]) else None,
                                qty=float(r["qty"]) if pd.notna(r["qty"]) else None,
                                unit_price_rmb=float(r["unit_price_rmb"]) if pd.notna(r["unit_price_rmb"]) else None,
                                line_total_rmb=float(r["line_total_rmb"]) if pd.notna(r["line_total_rmb"]) else None,
                            ))

                        db.commit()
                        st.success(f"已保存 ✅ PO={po_no}（items={len(items)}）")

                except Exception as e:
                    db.rollback()
                    st.error("保存失败 ❌")
                    st.exception(e)
                finally:
                    db.close()

        except Exception as e:
            st.error("解析失败 ❌")
            st.exception(e)

with tab2:
    st.subheader("查詢 PO")
    q = st.text_input("輸入 PO No（例如 PO-20260206-3）")
    if st.button("🔎 查詢"):
        db = get_db()
        try:
            h = db.get(POHeader, q.strip()) if q else None
            if not h:
                st.warning("找不到该 PO")
            else:
                st.json({
                    "po_no": h.po_no,
                    "supplier_name": h.supplier_name,
                    "order_date": h.order_date,
                })
                rows = db.query(POItem).filter(POItem.po_no == h.po_no).order_by(POItem.seq.asc()).all()
                df = pd.DataFrame([{
                    "seq": r.seq,
                    "product_code": r.product_code,
                    "product_name_cn": r.product_name_cn,
                    "qty": r.qty,
                    "unit_price_rmb": r.unit_price_rmb,
                    "ncm": r.ncm,
                    "product_name_pt": r.product_name_pt,
                    "line_total_rmb": r.line_total_rmb,
                } for r in rows])
                st.dataframe(df, use_container_width=True)
        finally:
            db.close()
