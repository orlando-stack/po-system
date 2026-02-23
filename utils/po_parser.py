import re
import pandas as pd
import openpyxl

def parse_po_excel(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    # 找 PO 号
    po_no = None
    supplier_name = None

    for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
        for cell in row:
            if cell.value:
                text = str(cell.value)
                if "买方合同号" in text or "PO" in text:
                    po_no = ws.cell(row=cell.row, column=cell.column + 1).value
                if "厂商名称" in text or "供应商" in text:
                    supplier_name = ws.cell(row=cell.row, column=cell.column + 1).value

    # 找明细表头
    header_row = None
    for r in range(1, 50):
        row_values = [str(ws.cell(r, c).value) for c in range(1, 20)]
        joined = " ".join(row_values)
        if "货号" in joined and ("数量" in joined or "订单数量" in joined):
            header_row = r
            break

    if not header_row:
        raise Exception("找不到明细表头")

    items = []

    for r in range(header_row + 1, header_row + 200):
        seq = ws.cell(r, 1).value
        if not seq:
            continue
        if not str(seq).isdigit():
            continue

        items.append({
            "seq": seq,
            "product_code": ws.cell(r, 2).value,
            "product_name_cn": ws.cell(r, 3).value,
            "qty": ws.cell(r, 8).value,
            "unit_price_rmb": ws.cell(r, 14).value,
            "ncm": ws.cell(r, 17).value,
            "product_name_pt": ws.cell(r, 18).value,
        })

    df = pd.DataFrame(items)

    header = {
        "po_no": po_no,
        "supplier_name": supplier_name
    }

    return header, df
